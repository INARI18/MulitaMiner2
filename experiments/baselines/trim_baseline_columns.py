"""Trim baseline XLSX columns to the scanner's own record fields.

The union layout leaves every scanner's baseline carrying foreign columns
(e.g. `instances` on OpenVAS) that are always empty and only ignored at eval
time. This drops them so a baseline matches its scanner's record model, which
is also cleaner as a training dataset.

Scanner is the file's parent-folder name (the resources/<scanner>/ convention).
Refuses to drop a non-empty foreign column.

Usage: uv run python tools/trim_baseline_columns.py resources/nessus/*.xlsx
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, "src")
from mulitaminer.scanner_engine import get_scanner  # noqa: E402
from mulitaminer.writers import columns_for  # noqa: E402


def trim(path: Path) -> None:
    scanner = path.parent.name
    keep = columns_for(get_scanner(scanner).record_type)

    ws = openpyxl.load_workbook(path).active
    cols = [c.value for c in ws[1]]
    rows = [dict(zip(cols, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    dropped = [c for c in cols if c not in keep]
    for c in dropped:
        if any(r.get(c) not in (None, "", "[]", "{}") for r in rows):
            raise SystemExit(f"{path.name}: refusing to drop non-empty column {c!r}")

    out = openpyxl.Workbook()
    o = out.active
    o.append(keep)
    for r in rows:
        o.append([r.get(c, "") for c in keep])
    out.save(path)
    print(f"{path.name}: {len(cols)} -> {len(keep)} cols (dropped {dropped})")


def main() -> None:
    for arg in sys.argv[1:]:
        trim(Path(arg))


if __name__ == "__main__":
    main()
